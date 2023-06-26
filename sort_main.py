"""
sort_main - 根据excel表的指定sheet进行排序

Author: hanayo
Date： 2023/6/20
"""
from copy import copy
import queue
from threading import Thread
import openpyxl
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import pathlib
from tkinter.filedialog import askopenfilename
import xlrd
from openpyxl.utils import column_index_from_string as col_index
import win32com.client as win32
import time


class SortSheetTool(ttk.Frame):

    queue = queue.Queue()

    def __init__(self, master: ttk.Window, canvas=None):
        super().__init__(master, padding=15)
        self.pack(fill=BOTH, expand=YES)
        if canvas:
            canvas.destroy()
        master.title("Excel给Sheet排序")
        self.sheet_list = []
        self.log = ""

        # 获取当前路径
        _path = pathlib.Path().absolute().as_posix()
        self.path_var = ttk.StringVar(value=_path)

        option_text = "请选择文件，然后点击按钮开始排序"
        self.option_frame = ttk.Labelframe(self, text=option_text, padding=15)
        self.option_frame.pack(fill=X, expand=YES, anchor=N)

        output_text = "运行结果如下："
        self.output = ttk.Labelframe(self, text=output_text, padding=15)
        self.output.pack(fill=X, expand=YES, anchor=N, pady=10)

        self.create_file_select()
        # 创建开始按钮那一行
        self.create_btn_row()
        # 创建保存结果那一行
        def_message = "excel排序工具ver1.0，请选择文件后点击开始排序，\n目前仅支持xls文件。"
        self.output_text = ttk.StringVar()
        self.output_text.set(def_message)
        self.create_output()  # 创建结果那一行

        self.start_time = time.time()

    def create_file_select(self):
        """选择文件那一行"""
        file_row = ttk.Frame(self.option_frame)
        file_row.pack(fill=X, expand=YES)
        file_lab = ttk.Label(file_row, text="请选择文件", width=10)
        file_lab.pack(side=LEFT, padx=(15, 0))
        file_entry = ttk.Entry(file_row, textvariable=self.path_var, width=50)
        file_entry.pack(side=LEFT, fill=X, expand=YES, padx=5)
        browse_btn = ttk.Button(
            master=file_row,
            text="浏览",
            command=self.choose_file,
            width=8
        )
        browse_btn.pack(side=LEFT, padx=5)

    def choose_file(self):
        """选择文件"""
        file = askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel Files", "*.xls"), ("Excel Files", "*.xlsx")]
        )
        if file:
            self.path_var.set(file)

    def create_btn_row(self):
        """创建按钮这一行"""
        btn_row = ttk.Frame(self.option_frame)
        btn_row.pack(fill=X, expand=YES, pady=15)
        # 这个label用来占位，主要是为了后面按钮的对齐
        path_lbl = ttk.Label(btn_row, text="", width=10)
        path_lbl.pack(side=LEFT, padx=(15, 0))
        # 第一个按钮
        quit_button = ttk.Button(
            master=btn_row,
            text="关闭",
            command=self.quit,
            style="success solid toolbutton",
            width=8
        )
        quit_button.pack(side=LEFT, padx=5)

        # 这个label用来站位，主要是为了后面按钮的对齐
        path_lbl2 = ttk.Label(btn_row, text="", width=50)
        path_lbl2.pack(side=LEFT, padx=(15, 0))
        # 第二个按钮
        st_button = ttk.Button(
            master=btn_row,
            text="开始排序",
            command=self.start_work,
            style=OUTLINE,
            width=8
        )
        st_button.pack(side=LEFT, padx=5)

    def start_work(self):
        """先生成xlsx文件，再建一个线程运行do_work"""
        self.start_time = time.time()
        self.log = "正在转换xlsx文件-->"
        self.output_text.set(self.log)
        self.update_idletasks()

        file = self.path_var.get()
        th1 = Thread(target=self.trans_xlsx, args=(file,))
        th1.start()
        th1.join()
        trans_log, is_ok = self.queue.get()

        self.log += trans_log
        cost_time = f"，已用时：{time.time() - self.start_time:.2f}秒。\n"
        self.log += cost_time
        self.output_text.set(self.log)
        self.update_idletasks()
        if is_ok:
            Thread(target=self.do_work, daemon=True).start()

    def do_work(self):
        """主要工作，先读取，再排序，再保存"""
        self.read_excel()
        self.sort_sheets()
        self.save_xlsx()

    @staticmethod
    def trans_xlsx(file):
        """将Excel转换成xlsx"""
        file_name = pathlib.Path(file).name
        save_path = pathlib.Path(file).parent
        # 创建 Excel Application 对象
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        try:
            wb = excel.Workbooks.Open(file)
            new_name = file_name + "x"
            wb.SaveAs(f"{save_path}/{new_name}", FileFormat=51)
            wb.Close(SaveChanges=1)
            txt_log = "转换成功"
            SortSheetTool.queue.put((txt_log, True))
            return txt_log, True
        except Exception as e:
            txt_log = f'转换失败！原因：{e}'
            SortSheetTool.queue.put((txt_log, False))
            return txt_log, False
        finally:
            # 退出 Excel
            excel.Application.Quit()

    def read_excel(self):
        """读取excel文件，用xlrd，读取效率高，"""
        self.log += "\n开始读取excel-->"
        self.output_text.set(self.log)
        try:
            workbook = xlrd.open_workbook(self.path_var.get())
            sheet_names = workbook.sheet_names()
            for sheet_name in sheet_names:
                if sheet_name == "XDO_METADATA":
                    pass
                else:
                    sheet_obj = {
                        "sheet_name": "",
                        "sheet_index": 0
                    }
                    sheet = workbook.sheet_by_name(sheet_name)
                    sheet_obj["sheet_name"] = sheet_name
                    sheet_obj["sheet_index"] = int(sheet.cell(5, col_index("Q") - 1).value)
                    self.sheet_list.append(sheet_obj)
            workbook.release_resources()
            # 关闭 Excel 文件
            self.log += f"读取完毕，共读取{len(self.sheet_list)}个sheet。\n"
            self.output_text.set(self.log)
            self.update_idletasks()
        except Exception as e:
            self.log += f"读取失败，错误原因如下：\n{e}"

    def sort_sheets(self):
        """对sheet进行排序"""
        self.log += "开始进行排序-->"
        self.output_text.set(self.log)
        sorted_list = sorted(self.sheet_list, key=lambda x: x["sheet_index"])
        cost_time = f"，已用时：{time.time() - self.start_time:.2f}秒。\n"
        self.sheet_list = sorted_list
        self.log += f"排序已完成{cost_time}\n即将开始复制-->"
        self.output_text.set(self.log)

    def save_xlsx(self):
        file_name = self.path_var.get() + "x"
        workbook = openpyxl.load_workbook(file_name)
        sheet_names = []
        for i in range(len(self.sheet_list)):
            sheet_names.append(self.sheet_list[i]["sheet_name"])
        new_wb = openpyxl.Workbook()
        count = 1
        self.log += "正在复制:\n"
        self.output_text.set(self.log)
        log_temp = self.log
        for sheet_name in sheet_names:
            self.update_idletasks()
            if count % 10 == 0:
                cost_time = f"，已用时：{time.time() - self.start_time:.2f}秒。\n"
                count_text = log_temp + f"正在复制第{count}个sheet-->{cost_time}"
                self.output_text.set(count_text)
            source_worksheet = workbook[sheet_name]
            new_worksheet = new_wb.create_sheet(sheet_name)

            # 复制格式与样式到新Sheet
            new_worksheet.sheet_format = source_worksheet.sheet_format
            new_worksheet.sheet_properties = source_worksheet.sheet_properties
            new_worksheet.merged_cells = source_worksheet.merged_cells
            new_worksheet.page_margins = source_worksheet.page_margins
            new_worksheet.page_setup = source_worksheet.page_setup
            new_worksheet.print_options = source_worksheet.print_options

            # 复制数据到新Sheet
            for row in source_worksheet.iter_rows(min_row=1, max_row=source_worksheet.max_row,
                                                  min_col=1, max_col=source_worksheet.max_column):
                for cell in row:
                    new_cell = new_worksheet.cell(row=cell.row, column=cell.column)
                    new_cell.value = cell.value
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)
            # 复制列宽和行高
            for i in range(1, source_worksheet.max_column + 1):
                new_worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = \
                    source_worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width
            for i in range(1, source_worksheet.max_row + 1):
                new_worksheet.row_dimensions[i].height = source_worksheet.row_dimensions[i].height
            count += 1
        self.log += f"已复制{count}个sheet，正在保存-->"
        self.output_text.set(self.log)
        # 删除默认第一个空的sheet
        blank_sheet = new_wb.worksheets[0]
        new_wb.remove(blank_sheet)
        save_path = pathlib.Path(self.path_var.get()).parent
        save_name = pathlib.Path(self.path_var.get()).name
        new_wb.save(f"{save_path}/已排序_{save_name}x")
        final_time = f"\n总用时：{time.time() - self.start_time:.2f}秒。"
        self.log += f"已完成。\n\n保存在{save_path}，\n文件名是：已排序_{save_name}x{final_time}"
        self.output_text.set(self.log)
        self.sheet_list = []

    def create_output(self):
        output_lb0 = ttk.Label(self.output, text="", width=10)
        output_lb0.pack(side=LEFT, padx=(15, 0))
        output_lb = ttk.Label(self.output,
                              textvariable=self.output_text,
                              width=50, style=INFO
                              )
        output_lb.pack(side=LEFT, padx=5)


if __name__ == '__main__':
    root = ttk.Window("sheet排序整理工具", "journal")
    SortSheetTool(root)
    root.mainloop()
