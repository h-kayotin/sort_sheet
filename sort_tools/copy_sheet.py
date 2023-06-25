"""
copy_sheet - 用openpyxl复制表格，附带格式

Author: hanayo
Date： 2023/6/25
"""

from copy import copy
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
import pathlib


def copy_sheet(file, save_path):
    """
    复制带格式的sheet，通过openpyxl
    :param file: 文件完整路径
    :param save_path: 保存路径
    :return: 无返回值
    """
    workbook = openpyxl.load_workbook(file)  # type:Workbook
    sheet_names = workbook.sheetnames

    new_wb = openpyxl.Workbook()
    for sheet_name in sheet_names:
        source_worksheet = workbook[sheet_name]  # type: Worksheet
        new_worksheet = new_wb.create_sheet(sheet_name)

        # 复制格式与样式到新Sheet
        new_worksheet.sheet_format = source_worksheet.sheet_format
        new_worksheet.sheet_properties = source_worksheet.sheet_properties
        new_worksheet.merged_cells = source_worksheet.merged_cells
        new_worksheet.page_margins = source_worksheet.page_margins
        new_worksheet.page_setup = source_worksheet.page_setup
        new_worksheet.print_options = source_worksheet.print_options

        # 复制数据到新Sheet
        for row in source_worksheet.iter_rows(min_row=1, max_row=source_worksheet.max_row,
                                              min_col=1, max_col=source_worksheet.max_column):
            for cell in row:
                new_cell = new_worksheet.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
        # 复制列宽和行高
        for i in range(1, source_worksheet.max_column + 1):
            new_worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = \
                source_worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width
        for i in range(1, source_worksheet.max_row + 1):
            new_worksheet.row_dimensions[i].height = source_worksheet.row_dimensions[i].height
    new_name = pathlib.Path(file).name
    new_wb.save(f"{save_path}/已复制_{new_name}")
